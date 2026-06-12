"""Tests para agente1/excel_writer.py."""

from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pytest

from agente1.excel_writer import COLUMNAS, EscritorCasos


@dataclass
class DatosReclamoFalso:
    """Doble de prueba con los 13 campos de DatosReclamo."""

    rut_cliente: str = "12.345.678-9"
    nombre_cliente: str = "María González Fuentes"
    numero_tarjeta: str = "4521"
    monto_reclamado: float = 185900
    moneda: str = "CLP"
    fecha_transaccion: str = "2026-05-20"
    fecha_reclamo: str = "2026-05-25"
    nombre_comercio: str = "TiendaOnline Express Ltda."
    canal_venta: str = "No presencial"
    tiene_3ds: str = "No"
    tipo_fraude: str = "Compra no reconocida por internet"
    descripcion_reclamo: str = "No reconozco esta compra"
    numero_operacion: str = "OP-2026-001-789456"


@dataclass
class DatosCorreoFalso:
    """Doble de prueba con ruta_original."""

    ruta_original: Path


@pytest.fixture
def escritor() -> EscritorCasos:
    return EscritorCasos()


def test_inicializar_excel_crea_archivo_con_encabezados(
    escritor: EscritorCasos, tmp_path: Path
) -> None:
    ruta = tmp_path / "output" / "casos.xlsx"

    escritor.inicializar_excel(ruta)

    assert ruta.exists()

    libro = openpyxl.load_workbook(ruta)
    encabezados = next(libro.active.iter_rows(values_only=True))
    assert list(encabezados) == COLUMNAS


def test_inicializar_excel_no_sobrescribe_si_existe(
    escritor: EscritorCasos, tmp_path: Path
) -> None:
    ruta = tmp_path / "casos.xlsx"
    escritor.inicializar_excel(ruta)

    libro = openpyxl.load_workbook(ruta)
    libro.active.append(["existente"] + [""] * (len(COLUMNAS) - 1))
    libro.save(ruta)

    escritor.inicializar_excel(ruta)

    libro = openpyxl.load_workbook(ruta)
    filas = list(libro.active.iter_rows(values_only=True))
    assert len(filas) == 2  # encabezado + fila existente


def test_escribir_caso_crea_archivo_y_registro(escritor: EscritorCasos, tmp_path: Path) -> None:
    ruta = tmp_path / "casos.xlsx"
    datos = DatosReclamoFalso()
    correo = DatosCorreoFalso(ruta_original=Path("input/emails/caso_001.eml"))

    id_caso = escritor.escribir_caso(
        datos=datos,
        correo=correo,
        hash_dedup="hash-abc",
        estado="PROCESADO",
        ruta_excel=ruta,
    )

    libro = openpyxl.load_workbook(ruta)
    filas = list(libro.active.iter_rows(values_only=True))

    assert len(filas) == 2
    encabezados, fila = filas

    registro = dict(zip(encabezados, fila))
    assert registro["id_caso"] == id_caso
    assert registro["archivo_origen"] == "caso_001.eml"
    assert registro["hash_dedup"] == "hash-abc"
    assert registro["rut_cliente"] == "12.345.678-9"
    assert registro["monto_reclamado"] == 185900
    assert registro["estado"] == "PROCESADO"
    assert registro["id_caso_original"] in ("", None)


def test_escribir_caso_duplicado_con_id_original(escritor: EscritorCasos, tmp_path: Path) -> None:
    ruta = tmp_path / "casos.xlsx"
    datos = DatosReclamoFalso()
    correo = DatosCorreoFalso(ruta_original=Path("input/emails/caso_002.eml"))

    id_caso = escritor.escribir_caso(
        datos=datos,
        correo=correo,
        hash_dedup="hash-dup",
        estado="DUPLICADO",
        ruta_excel=ruta,
        id_caso_original="caso-original-123",
        notas="Duplicado del caso original",
    )

    libro = openpyxl.load_workbook(ruta)
    _, fila = list(libro.active.iter_rows(values_only=True))
    registro = dict(zip(COLUMNAS, fila))

    assert registro["estado"] == "DUPLICADO"
    assert registro["id_caso_original"] == "caso-original-123"
    assert registro["notas"] == "Duplicado del caso original"
    assert registro["id_caso"] == id_caso


def test_escribir_caso_multiples_registros_se_acumulan(
    escritor: EscritorCasos, tmp_path: Path
) -> None:
    ruta = tmp_path / "casos.xlsx"
    correo1 = DatosCorreoFalso(ruta_original=Path("input/emails/caso_001.eml"))
    correo2 = DatosCorreoFalso(ruta_original=Path("input/emails/caso_002.eml"))

    escritor.escribir_caso(DatosReclamoFalso(), correo1, "hash-1", "PROCESADO", ruta)
    escritor.escribir_caso(DatosReclamoFalso(), correo2, "hash-2", "PROCESADO", ruta)

    libro = openpyxl.load_workbook(ruta)
    filas = list(libro.active.iter_rows(values_only=True))

    assert len(filas) == 3  # encabezado + 2 registros


def test_escribir_caso_genera_uuid_distinto_por_caso(
    escritor: EscritorCasos, tmp_path: Path
) -> None:
    ruta = tmp_path / "casos.xlsx"
    correo = DatosCorreoFalso(ruta_original=Path("input/emails/caso_001.eml"))

    id_caso_1 = escritor.escribir_caso(DatosReclamoFalso(), correo, "hash-1", "PROCESADO", ruta)
    id_caso_2 = escritor.escribir_caso(DatosReclamoFalso(), correo, "hash-2", "PROCESADO", ruta)

    assert id_caso_1 != id_caso_2
