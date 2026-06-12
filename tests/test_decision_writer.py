"""Tests para agente2/decision_writer.py."""

from pathlib import Path

import openpyxl
import pytest

from agente2.caso_reader import CasoFraude
from agente2.decision_writer import COLUMNAS_DECISIONES, EscritorDecisiones
from agente2.regla_3ds import VERSION_REGLA, DecisionChargeback


def _caso(**overrides) -> CasoFraude:
    valores = {
        "id_caso": "caso-001",
        "fecha_procesamiento": "2026-05-25T10:00:00",
        "archivo_origen": "caso_001.eml",
        "hash_dedup": "hash-1",
        "rut_cliente": "12.345.678-9",
        "nombre_cliente": "María González Fuentes",
        "numero_tarjeta": "4521",
        "monto_reclamado": 185900,
        "moneda": "CLP",
        "fecha_transaccion": "2026-05-20",
        "fecha_reclamo": "2026-05-25",
        "nombre_comercio": "TiendaOnline Express Ltda.",
        "canal_venta": "No presencial",
        "tiene_3ds": "No",
        "tipo_fraude": "Compra no reconocida por internet",
        "descripcion_reclamo": "No reconozco esta compra",
        "numero_operacion": "OP-2026-001-789456",
        "estado": "PROCESADO",
        "id_caso_original": "",
        "notas": "",
    }
    valores.update(overrides)
    return CasoFraude(**valores)


def _decision(**overrides) -> DecisionChargeback:
    valores = {
        "aplica_chargeback": "Sí",
        "justificacion": "Comercio sin 3DS en venta no presencial.",
        "codigo_razon": "Visa 10.4 / MC 4837",
        "estado_decision": "RESUELTA",
        "version_regla": VERSION_REGLA,
        "dias_entre_transaccion_y_reclamo": 5,
        "alertas": "",
    }
    valores.update(overrides)
    return DecisionChargeback(**valores)


@pytest.fixture
def escritor() -> EscritorDecisiones:
    return EscritorDecisiones()


def test_inicializar_excel_crea_encabezados(escritor: EscritorDecisiones, tmp_path: Path) -> None:
    ruta = tmp_path / "decisiones.xlsx"

    escritor.inicializar_excel(ruta)

    libro = openpyxl.load_workbook(ruta)
    encabezados = next(libro.active.iter_rows(values_only=True))
    assert list(encabezados) == COLUMNAS_DECISIONES


def test_inicializar_excel_no_sobrescribe(escritor: EscritorDecisiones, tmp_path: Path) -> None:
    ruta = tmp_path / "decisiones.xlsx"
    escritor.inicializar_excel(ruta)

    libro = openpyxl.load_workbook(ruta)
    libro.active.append(["existente"] + [""] * (len(COLUMNAS_DECISIONES) - 1))
    libro.save(ruta)

    escritor.inicializar_excel(ruta)

    libro = openpyxl.load_workbook(ruta)
    assert len(list(libro.active.iter_rows(values_only=True))) == 2


def test_escribir_decision_chargeback_si(escritor: EscritorDecisiones, tmp_path: Path) -> None:
    ruta = tmp_path / "decisiones.xlsx"
    caso = _caso()
    decision = _decision()

    id_decision = escritor.escribir_decision(caso, decision, ruta)

    libro = openpyxl.load_workbook(ruta)
    encabezados, fila = list(libro.active.iter_rows(values_only=True))
    registro = dict(zip(encabezados, fila))

    assert registro["id_decision"] == id_decision
    assert registro["id_caso"] == "caso-001"
    assert registro["rut_cliente"] == "12.345.678-9"
    assert registro["monto_reclamado"] == 185900
    assert registro["nombre_comercio"] == "TiendaOnline Express Ltda."
    assert registro["canal_venta"] == "No presencial"
    assert registro["tiene_3ds"] == "No"
    assert registro["aplica_chargeback"] == "Sí"
    assert registro["codigo_razon"] == "Visa 10.4 / MC 4837"
    assert registro["version_regla"] == VERSION_REGLA
    assert registro["dias_entre_transaccion_y_reclamo"] == 5
    assert registro["estado_decision"] == "RESUELTA"


def test_escribir_decision_pendiente_revision(escritor: EscritorDecisiones, tmp_path: Path) -> None:
    ruta = tmp_path / "decisiones.xlsx"
    caso = _caso(canal_venta="Presencial")
    decision = _decision(
        aplica_chargeback="—",
        justificacion="Regla 3DS no aplicable a transacciones presenciales.",
        codigo_razon="",
        estado_decision="PENDIENTE_REVISION",
        alertas="FUERA_DE_PLAZO",
    )

    escritor.escribir_decision(caso, decision, ruta)

    libro = openpyxl.load_workbook(ruta)
    encabezados, fila = list(libro.active.iter_rows(values_only=True))
    registro = dict(zip(encabezados, fila))

    assert registro["aplica_chargeback"] == "—"
    assert registro["codigo_razon"] in ("", None)
    assert registro["alertas"] == "FUERA_DE_PLAZO"
    assert registro["estado_decision"] == "PENDIENTE_REVISION"


def test_escribir_multiples_decisiones_se_acumulan(escritor: EscritorDecisiones, tmp_path: Path) -> None:
    ruta = tmp_path / "decisiones.xlsx"

    id_1 = escritor.escribir_decision(_caso(id_caso="caso-001"), _decision(), ruta)
    id_2 = escritor.escribir_decision(
        _caso(id_caso="caso-002"), _decision(aplica_chargeback="No"), ruta
    )

    libro = openpyxl.load_workbook(ruta)
    filas = list(libro.active.iter_rows(values_only=True))

    assert len(filas) == 3  # encabezado + 2
    assert id_1 != id_2
