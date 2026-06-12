"""Tests para agente1/dedup.py."""

from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pytest

from agente1.dedup import Deduplicador


@dataclass
class DatosReclamoFalso:
    """Doble de prueba con los campos mínimos requeridos para dedup."""

    rut_cliente: str
    numero_operacion: str
    monto_reclamado: float


@pytest.fixture
def deduplicador() -> Deduplicador:
    return Deduplicador()


# === calcular_hash ===

def test_calcular_hash_es_deterministico(deduplicador: Deduplicador) -> None:
    datos = DatosReclamoFalso("12.345.678-9", "OP-2026-001-789456", 185900)

    hash1 = deduplicador.calcular_hash(datos)
    hash2 = deduplicador.calcular_hash(datos)

    assert hash1 == hash2
    assert len(hash1) == 64  # sha256 hex


def test_calcular_hash_int_y_float_iguales(deduplicador: Deduplicador) -> None:
    datos_int = DatosReclamoFalso("12.345.678-9", "OP-1", 185900)
    datos_float = DatosReclamoFalso("12.345.678-9", "OP-1", 185900.0)

    assert deduplicador.calcular_hash(datos_int) == deduplicador.calcular_hash(datos_float)


def test_calcular_hash_difiere_si_cambia_un_campo(deduplicador: Deduplicador) -> None:
    base = DatosReclamoFalso("12.345.678-9", "OP-1", 100000)
    distinto_rut = DatosReclamoFalso("11.111.111-1", "OP-1", 100000)
    distinto_monto = DatosReclamoFalso("12.345.678-9", "OP-1", 200000)

    hash_base = deduplicador.calcular_hash(base)

    assert hash_base != deduplicador.calcular_hash(distinto_rut)
    assert hash_base != deduplicador.calcular_hash(distinto_monto)


def test_calcular_hash_numero_operacion_vacio(deduplicador: Deduplicador) -> None:
    datos = DatosReclamoFalso("12.345.678-9", "", 100000)

    hash_dedup = deduplicador.calcular_hash(datos)

    assert len(hash_dedup) == 64


# === existe_en_excel ===

def test_existe_en_excel_archivo_no_existe(deduplicador: Deduplicador, tmp_path: Path) -> None:
    resultado = deduplicador.existe_en_excel("hash123", tmp_path / "casos.xlsx")

    assert resultado is None


def _crear_excel_casos(ruta: Path, filas: list[tuple]) -> None:
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.append(["id_caso", "fecha_procesamiento", "hash_dedup", "estado"])
    for fila in filas:
        hoja.append(fila)
    libro.save(ruta)


def test_existe_en_excel_hash_encontrado(deduplicador: Deduplicador, tmp_path: Path) -> None:
    ruta = tmp_path / "casos.xlsx"
    _crear_excel_casos(
        ruta,
        [
            ("caso-001", "2026-06-10T10:00:00", "hash-abc", "PROCESADO"),
            ("caso-002", "2026-06-11T10:00:00", "hash-xyz", "PROCESADO"),
        ],
    )

    resultado = deduplicador.existe_en_excel("hash-xyz", ruta)

    assert resultado == "caso-002"


def test_existe_en_excel_hash_no_encontrado(deduplicador: Deduplicador, tmp_path: Path) -> None:
    ruta = tmp_path / "casos.xlsx"
    _crear_excel_casos(ruta, [("caso-001", "2026-06-10T10:00:00", "hash-abc", "PROCESADO")])

    resultado = deduplicador.existe_en_excel("hash-no-existe", ruta)

    assert resultado is None


def test_existe_en_excel_archivo_vacio(deduplicador: Deduplicador, tmp_path: Path) -> None:
    ruta = tmp_path / "casos.xlsx"
    libro = openpyxl.Workbook()
    libro.save(ruta)

    resultado = deduplicador.existe_en_excel("hash-abc", ruta)

    assert resultado is None


def test_existe_en_excel_columnas_faltantes(deduplicador: Deduplicador, tmp_path: Path) -> None:
    ruta = tmp_path / "casos.xlsx"
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.append(["columna_a", "columna_b"])
    hoja.append(["x", "y"])
    libro.save(ruta)

    resultado = deduplicador.existe_en_excel("hash-abc", ruta)

    assert resultado is None
